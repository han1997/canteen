from __future__ import annotations

from datetime import datetime
from pathlib import Path
from random import randint

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from sqlalchemy import Select, select
from sqlalchemy.orm import Session

from app.models import ExportJob, MealSlot, Order, OrderItem, User


BACKEND_ROOT = Path(__file__).resolve().parents[2]


def build_job_no() -> str:
    return f"EX{datetime.utcnow().strftime('%Y%m%d%H%M%S')}{randint(1000, 9999)}"


def create_export_job(
    db: Session,
    request_user_id: int,
    from_date,
    to_date,
    meal_type: str,
    meal_category: str,
) -> ExportJob:
    job = ExportJob(
        job_no=build_job_no(),
        request_user_id=request_user_id,
        from_date=from_date,
        to_date=to_date,
        meal_type=meal_type,
        meal_category=meal_category,
        status="queued",
    )
    db.add(job)
    db.flush()
    return job


def _build_query(job: ExportJob) -> Select:
    stmt = (
        select(
            Order.id.label("order_id"),
            Order.order_no,
            User.police_no,
            User.real_name,
            MealSlot.meal_date,
            MealSlot.meal_type,
            Order.meal_category,
            Order.status,
            Order.booked_at,
            OrderItem.item_name,
            OrderItem.quantity,
            OrderItem.unit,
            OrderItem.unit_price,
        )
        .join(User, User.id == Order.user_id)
        .join(MealSlot, MealSlot.id == Order.slot_id)
        .outerjoin(OrderItem, OrderItem.order_id == Order.id)
        .where(
            MealSlot.meal_date >= job.from_date,
            MealSlot.meal_date <= job.to_date,
        )
    )

    if job.meal_type != "all":
        stmt = stmt.where(MealSlot.meal_type == job.meal_type)
    if job.meal_category != "all":
        stmt = stmt.where(Order.meal_category == job.meal_category)

    return stmt.order_by(MealSlot.meal_date, MealSlot.meal_type, Order.booked_at, OrderItem.id)


def _format_sheet(ws) -> None:
    ws.freeze_panes = "A2"
    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 10
    ws.column_dimensions["D"].width = 14

    header_fill = PatternFill(fill_type="solid", start_color="0B2A4A", end_color="0B2A4A")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    thin = Side(style="thin", color="D8E2EF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=4):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical="center")


DATE_BANNER_FILL = PatternFill(fill_type="solid", start_color="E8F1FF", end_color="E8F1FF")
DATE_BANNER_FONT = Font(color="0B2A4A", bold=True, size=11)


def _format_item_sheet(ws) -> None:
    ws.freeze_panes = "A2"
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 10

    header_fill = PatternFill(fill_type="solid", start_color="0B2A4A", end_color="0B2A4A")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    thin = Side(style="thin", color="D8E2EF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=3):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical="center")


def _group_rows(rows) -> list[dict]:
    grouped: list[dict] = []
    index_map: dict[int, int] = {}

    for row in rows:
        order_id = int(row.order_id)
        if order_id not in index_map:
            index_map[order_id] = len(grouped)
            grouped.append(
                {
                    "order_id": order_id,
                    "order_no": row.order_no,
                    "police_no": row.police_no,
                    "real_name": row.real_name,
                    "meal_date": row.meal_date,
                    "meal_type": row.meal_type,
                    "meal_category": row.meal_category,
                    "status": row.status,
                    "booked_at": row.booked_at,
                    "items": [],
                }
            )

        if row.item_name:
            grouped[index_map[order_id]]["items"].append(
                {
                    "item_name": row.item_name,
                    "quantity": float(row.quantity or 0),
                    "unit": row.unit or "份",
                    "unit_price": float(row.unit_price or 0),
                }
            )

    return grouped


def run_export_job(db: Session, job: ExportJob) -> ExportJob:
    export_dir = BACKEND_ROOT / "exports"
    export_dir.mkdir(parents=True, exist_ok=True)

    job.status = "running"
    db.flush()

    rows = db.execute(_build_query(job)).all()
    grouped_orders = _group_rows(rows)

    wb = Workbook()
    ws = wb.active
    ws.title = "订餐明细"
    ws.append(["姓名", "菜品", "份数", "价格(元)"])

    orders_by_date: dict[str, list[dict]] = {}
    for order in grouped_orders:
        date_text = order["meal_date"].strftime("%Y-%m-%d")
        orders_by_date.setdefault(date_text, []).append(order)

    banner_rows: list[tuple[int, str]] = []
    current_row = 2
    for date_text in sorted(orders_by_date.keys()):
        ws.append([date_text, "", "", ""])
        banner_rows.append((current_row, date_text))
        current_row += 1

        for order in orders_by_date[date_text]:
            items = order["items"] or [{"item_name": "-", "quantity": 0.0, "unit": "份", "unit_price": 0.0}]
            for item in items:
                subtotal = round(item["quantity"] * item["unit_price"], 2)
                ws.append(
                    [
                        order["real_name"],
                        item["item_name"],
                        item["quantity"],
                        subtotal,
                    ]
                )
                current_row += 1

    _format_sheet(ws)

    thin = Side(style="thin", color="D8E2EF")
    banner_border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for banner_row, date_text in banner_rows:
        ws.merge_cells(start_row=banner_row, start_column=1, end_row=banner_row, end_column=4)
        banner_cell = ws.cell(row=banner_row, column=1)
        banner_cell.value = date_text
        banner_cell.fill = DATE_BANNER_FILL
        banner_cell.font = DATE_BANNER_FONT
        banner_cell.alignment = Alignment(horizontal="center", vertical="center")
        for col_idx in range(1, 5):
            ws.cell(row=banner_row, column=col_idx).border = banner_border

    for col in ("C", "D"):
        for cell in ws[col][1:]:
            if isinstance(cell.value, (int, float)):
                cell.number_format = "0.00"

    items_sheet = wb.create_sheet("菜品订购人")
    items_sheet.append(["菜品", "姓名", "份数"])

    item_to_buyers: dict[str, list[dict]] = {}
    for order in grouped_orders:
        for item in order["items"]:
            item_to_buyers.setdefault(item["item_name"], []).append(
                {
                    "real_name": order["real_name"],
                    "quantity": item["quantity"],
                    "meal_date": order["meal_date"],
                }
            )

    item_row = 2
    for item_name in sorted(item_to_buyers.keys()):
        buyers = sorted(
            item_to_buyers[item_name],
            key=lambda b: (b["meal_date"], b["real_name"]),
        )
        start_row = item_row
        for buyer in buyers:
            items_sheet.append([item_name, buyer["real_name"], buyer["quantity"]])
            item_row += 1
        end_row = item_row - 1
        if end_row > start_row:
            items_sheet.merge_cells(start_row=start_row, start_column=1, end_row=end_row, end_column=1)
            items_sheet.cell(row=start_row, column=1).alignment = Alignment(
                horizontal="center", vertical="center"
            )

    _format_item_sheet(items_sheet)
    for cell in items_sheet["C"][1:]:
        if isinstance(cell.value, (int, float)):
            cell.number_format = "0.00"

    file_name = f"订餐统计_{job.from_date.strftime('%Y%m%d')}_{job.to_date.strftime('%Y%m%d')}_{job.job_no}.xlsx"
    output = (export_dir / file_name).resolve()
    wb.save(output)

    job.status = "done"
    job.file_path = str(output)
    return job
