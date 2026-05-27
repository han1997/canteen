import re
from datetime import date, datetime, time
from io import BytesIO
from pathlib import Path
from random import randint
from typing import Annotated
from uuid import uuid4

from fastapi import APIRouter, Depends, File, HTTPException, Query, Request, UploadFile
from openpyxl import load_workbook
from sqlalchemy import case, func, or_, select
from sqlalchemy.exc import IntegrityError
from sqlalchemy.orm import Session, joinedload

from app.core.security import get_current_user, get_password_hash, role_required
from app.core.config import settings
from app.db.session import get_db
from app.models import (
    MealCategoryEnum,
    MealPackage,
    MealPackageItem,
    MealPackageMealType,
    MealSlot,
    MealTypeEnum,
    Order,
    OrderItem,
    OrderStatusEnum,
    RoleEnum,
    User,
    UserStatusEnum,
)
from app.schemas.admin import (
    AdminBulkImportResult,
    AdminUserCreateRequest,
    AdminUserOut,
    AdminUserRoleUpdateRequest,
    AdminUserStatusUpdateRequest,
    AdminUserUpdateRequest,
)
from app.schemas.admin_meal import (
    AdminMealPackageCreateRequest,
    AdminMealPackageOut,
    AdminMealPackageUpdateRequest,
    AdminMealSlotCreateRequest,
    AdminMealSlotOut,
    AdminMealSlotStatusUpdateRequest,
)
from app.schemas.common import ApiMessage
from app.services.audit_service import write_audit


router = APIRouter(prefix="/admin", tags=["admin"])

USER_MANAGE_DEPS = [Depends(role_required(RoleEnum.ADMIN, RoleEnum.SUPER_ADMIN))]
MEAL_MANAGE_DEPS = [Depends(role_required(RoleEnum.KITCHEN, RoleEnum.ADMIN, RoleEnum.SUPER_ADMIN))]


class DashboardStats:
    def __init__(
        self,
        target_date: date,
        breakfast_count: int,
        lunch_count: int,
        dinner_count: int,
        breakfast_open: bool,
        lunch_open: bool,
        dinner_open: bool,
    ):
        self.target_date = target_date
        self.breakfast_count = breakfast_count
        self.lunch_count = lunch_count
        self.dinner_count = dinner_count
        self.breakfast_open = breakfast_open
        self.lunch_open = lunch_open
        self.dinner_open = dinner_open


def _build_meal_package_out(pkg: MealPackage) -> AdminMealPackageOut:
    items = [
        {
            "id": item.id,
            "item_name": item.item_name,
            "quantity": float(item.quantity),
            "unit": item.unit,
            "item_type": item.item_type,
        }
        for item in sorted(pkg.items, key=lambda x: x.sort_order)
    ]
    return AdminMealPackageOut(
        id=pkg.id,
        meal_types=[mt.value for mt in pkg.meal_types],  # 改为列表
        package_code=pkg.package_code,
        package_name=pkg.package_name,
        meal_category=pkg.meal_category.value,
        is_selectable=pkg.is_selectable,
        image_url=pkg.image_url or "",
        price=float(pkg.price) if pkg.price is not None else None,
        unit=pkg.unit or "份",
        calories=pkg.calories,
        protein_g=float(pkg.protein_g) if pkg.protein_g is not None else None,
        carbs_g=float(pkg.carbs_g) if pkg.carbs_g is not None else None,
        fat_g=float(pkg.fat_g) if pkg.fat_g is not None else None,
        sort_order=pkg.sort_order,
        items=items,
    )


def _generate_package_code(meal_type: str | None = None) -> str:
    # 使用通用前缀 MP（Meal Package），不再依赖单一餐别
    return f"MP{uuid4().hex[:8].upper()}"


def _load_package_with_relations(db: Session, package_id: int) -> MealPackage | None:
    """加载菜品并预加载 items 和 meal_type_associations，避免懒加载 N+1。"""
    return db.scalar(
        select(MealPackage)
        .options(
            joinedload(MealPackage.items),
            joinedload(MealPackage.meal_type_associations),
        )
        .where(MealPackage.id == package_id)
    )


def _validate_police_no(police_no: str | None) -> bool:
    """验证警号格式：非空且长度在2-32之间"""
    if not police_no:
        return True  # None is valid (optional field)
    return 2 <= len(police_no) <= 32


def _validate_mobile(mobile: str | None) -> bool:
    """验证手机号格式：11位数字"""
    if not mobile:
        return True  # None is valid (optional field)
    return bool(re.match(r"^1\d{10}$", mobile))


@router.get("/users", response_model=list[AdminUserOut], dependencies=USER_MANAGE_DEPS)
def list_users(
    current_user: Annotated[User, Depends(get_current_user)],
    db: Annotated[Session, Depends(get_db)],
    keyword: str | None = Query(default=None),
):
    _ = current_user
    stmt = select(User)
    if keyword:
        stmt = stmt.where(
            (User.police_no.like(f"%{keyword}%"))
            | (User.real_name.like(f"%{keyword}%"))
            | (User.mobile.like(f"%{keyword}%"))
        )
    return db.scalars(stmt.order_by(User.id.desc())).all()


@router.post("/users", response_model=AdminUserOut, dependencies=USER_MANAGE_DEPS)
def create_user(
    payload: AdminUserCreateRequest,
    request: Request,
    current_user: Annotated[User, Depends(get_current_user)],
    db: Annotated[Session, Depends(get_db)],
):
    if payload.police_no:
        if db.scalar(select(User).where(User.police_no == payload.police_no)):
            raise HTTPException(status_code=400, detail="警号已存在")
    if payload.mobile:
        if db.scalar(select(User).where(User.mobile == payload.mobile)):
            raise HTTPException(status_code=400, detail="手机号已存在")

    user = User(
        police_no=payload.police_no,
        real_name=payload.real_name,
        dept_name=payload.dept_name,
        mobile=payload.mobile,
        role=RoleEnum(payload.role),
        status=UserStatusEnum.ACTIVE,
        password_hash=get_password_hash(payload.init_password),
    )
    db.add(user)
    db.flush()

    write_audit(
        db,
        actor_user_id=current_user.id,
        action="CREATE_USER",
        target_type="user",
        target_id=str(user.id),
        request_ip=request.client.host if request.client else None,
    )
    db.commit()
    db.refresh(user)
    return user


@router.post("/users/bulk-import", response_model=AdminBulkImportResult, dependencies=USER_MANAGE_DEPS)
async def bulk_import_users(
    file: UploadFile = File(...),
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """批量导入就餐人员。Excel 表头：警号 / 姓名 / 手机号（警号与手机号至少填一个）。
    默认密码 123456，角色 officer，部门 祁门县公安局。
    限制：文件最大 10MB，最多 1000 行数据。"""
    if not file.filename or not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail="仅支持 .xlsx 或 .xls 文件")

    content = await file.read()
    if len(content) > settings.bulk_import_max_file_size:
        raise HTTPException(
            status_code=400,
            detail=f"文件过大，最大支持 {settings.bulk_import_max_file_size // (1024 * 1024)}MB",
        )

    try:
        wb = load_workbook(BytesIO(content), read_only=True, data_only=True)
    except Exception:
        raise HTTPException(status_code=400, detail="无效的 Excel 文件")
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise HTTPException(status_code=400, detail="Excel 文件为空")

    if len(rows) > settings.bulk_import_max_rows + 1:  # +1 for header
        raise HTTPException(
            status_code=400,
            detail=f"数据行数过多，最多支持 {settings.bulk_import_max_rows} 行",
        )

    header = [str(c or "").strip() for c in rows[0]]
    police_no_col = next((i for i, h in enumerate(header) if "警号" in h), None)
    real_name_col = next((i for i, h in enumerate(header) if "姓名" in h), None)
    mobile_col = next((i for i, h in enumerate(header) if "手机" in h), None)

    if real_name_col is None:
        raise HTTPException(status_code=400, detail="表头缺少「姓名」列")

    # Pre-fetch existing police_nos and mobiles to avoid N+1 queries
    existing_police_nos = set(
        pno for pno in db.scalars(select(User.police_no).where(User.police_no.isnot(None))).all()
    )
    existing_mobiles = set(
        mob for mob in db.scalars(select(User.mobile).where(User.mobile.isnot(None))).all()
    )

    created = 0
    skipped = 0
    errors = []

    try:
        for idx, row in enumerate(rows[1:], start=2):
            if not row or all(c is None or str(c).strip() == "" for c in row):
                continue

            police_no = str(row[police_no_col]).strip() if police_no_col is not None and row[police_no_col] else None
            real_name = str(row[real_name_col]).strip() if row[real_name_col] else None
            mobile = str(row[mobile_col]).strip() if mobile_col is not None and row[mobile_col] else None

            # Validate non-empty after strip
            if police_no and not police_no:
                police_no = None
            if mobile and not mobile:
                mobile = None

            if not real_name:
                errors.append(f"第 {idx} 行：姓名为空")
                continue
            if not police_no and not mobile:
                errors.append(f"第 {idx} 行：警号与手机号至少填一个")
                continue

            # Validate format
            if not _validate_police_no(police_no):
                errors.append(f"第 {idx} 行：警号格式无效（长度应为 2-32）")
                continue
            if not _validate_mobile(mobile):
                errors.append(f"第 {idx} 行：手机号格式无效（应为 11 位数字）")
                continue

            # Check duplicates
            if police_no and police_no in existing_police_nos:
                skipped += 1
                continue
            if mobile and mobile in existing_mobiles:
                skipped += 1
                continue

            user = User(
                police_no=police_no,
                real_name=real_name,
                mobile=mobile,
                dept_name="祁门县公安局",
                role=RoleEnum.OFFICER,
                status=UserStatusEnum.ACTIVE,
                password_hash=get_password_hash("123456"),
            )
            db.add(user)
            if police_no:
                existing_police_nos.add(police_no)
            if mobile:
                existing_mobiles.add(mobile)
            created += 1

        db.commit()
    except IntegrityError as e:
        db.rollback()
        error_msg = str(e.orig) if e.orig else str(e)
        if "police_no" in error_msg or "uk_users_police_no" in error_msg:
            raise HTTPException(status_code=409, detail="批量导入失败：存在重复的警号")
        elif "mobile" in error_msg or "uk_users_mobile" in error_msg:
            raise HTTPException(status_code=409, detail="批量导入失败：存在重复的手机号")
        else:
            raise HTTPException(status_code=409, detail=f"批量导入失败：数据冲突 - {error_msg}")
    except ValueError as e:
        db.rollback()
        raise HTTPException(status_code=400, detail=f"批量导入失败：数据格式错误 - {str(e)}")
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"批量导入失败：{str(e)}")

    write_audit(
        db,
        actor_user_id=current_user.id,
        action="BULK_IMPORT_USERS",
        target_type="user",
        target_id="bulk",
        detail_json={"created": created, "skipped": skipped, "error_count": len(errors)},
    )
    db.commit()

    return AdminBulkImportResult(created=created, skipped=skipped, errors=errors)


@router.patch("/users/{user_id}", response_model=AdminUserOut, dependencies=USER_MANAGE_DEPS)
def update_user(
    user_id: int,
    payload: AdminUserUpdateRequest,
    request: Request,
    current_user: Annotated[User, Depends(get_current_user)],
    db: Annotated[Session, Depends(get_db)],
):
    """编辑用户基础信息：警号、姓名、部门、手机号。
    警号与手机号必须至少保留一个非空。"""
    user = db.get(User, user_id)
    if not user:
        raise HTTPException(status_code=404, detail="用户不存在")

    changes: dict = {}

    # 计算更新后的 police_no 和 mobile，校验至少保留一个
    new_police_no = payload.police_no if payload.police_no is not None else user.police_no
    new_mobile = payload.mobile if payload.mobile is not None else user.mobile
    if not new_police_no and not new_mobile:
        raise HTTPException(status_code=400, detail="警号与手机号至少保留一个")

    # 警号唯一性校验
    if payload.police_no is not None and payload.police_no != user.police_no:
        if payload.police_no:
            exists = db.scalar(
                select(User).where(User.police_no == payload.police_no, User.id != user_id)
            )
            if exists:
                raise HTTPException(status_code=400, detail="警号已被其他用户占用")
        changes["police_no"] = {"old": user.police_no, "new": payload.police_no}
        user.police_no = payload.police_no

    # 手机号唯一性校验
    if payload.mobile is not None and payload.mobile != user.mobile:
        if payload.mobile:
            exists = db.scalar(
                select(User).where(User.mobile == payload.mobile, User.id != user_id)
            )
            if exists:
                raise HTTPException(status_code=400, detail="手机号已被其他用户占用")
        changes["mobile"] = {"old": user.mobile, "new": payload.mobile}
        user.mobile = payload.mobile

    # 姓名
    if payload.real_name is not None and payload.real_name != user.real_name:
        if not payload.real_name:
            raise HTTPException(status_code=400, detail="姓名不能为空")
        changes["real_name"] = {"old": user.real_name, "new": payload.real_name}
        user.real_name = payload.real_name

    # 部门
    if payload.dept_name is not None and payload.dept_name != user.dept_name:
        if not payload.dept_name:
            raise HTTPException(status_code=400, detail="部门不能为空")
        changes["dept_name"] = {"old": user.dept_name, "new": payload.dept_name}
        user.dept_name = payload.dept_name

    if not changes:
        return user

    try:
        write_audit(
            db,
            actor_user_id=current_user.id,
            action="UPDATE_USER_INFO",
            target_type="user",
            target_id=str(user.id),
            request_ip=request.client.host if request.client else None,
            detail_json=changes,
        )
        db.commit()
    except IntegrityError as e:
        db.rollback()
        error_msg = str(e.orig) if e.orig else str(e)
        if "police_no" in error_msg:
            raise HTTPException(status_code=409, detail="警号已被占用")
        if "mobile" in error_msg:
            raise HTTPException(status_code=409, detail="手机号已被占用")
        raise HTTPException(status_code=409, detail=f"数据冲突：{error_msg}")

    db.refresh(user)
    return user


@router.patch("/users/{user_id}/role", response_model=ApiMessage, dependencies=USER_MANAGE_DEPS)
def update_role(
    user_id: int,
    payload: AdminUserRoleUpdateRequest,
    request: Request,
    current_user: Annotated[User, Depends(get_current_user)],
    db: Annotated[Session, Depends(get_db)],
):
    user = db.get(User, user_id)
    if not user:
        raise HTTPException(status_code=404, detail="用户不存在")

    user.role = RoleEnum(payload.role)
    write_audit(
        db,
        actor_user_id=current_user.id,
        action="UPDATE_USER_ROLE",
        target_type="user",
        target_id=str(user.id),
        request_ip=request.client.host if request.client else None,
        detail_json={"new_role": payload.role},
    )
    db.commit()
    return ApiMessage(message="角色已更新")


@router.patch("/users/{user_id}/status", response_model=ApiMessage, dependencies=USER_MANAGE_DEPS)
def update_status(
    user_id: int,
    payload: AdminUserStatusUpdateRequest,
    request: Request,
    current_user: Annotated[User, Depends(get_current_user)],
    db: Annotated[Session, Depends(get_db)],
):
    user = db.get(User, user_id)
    if not user:
        raise HTTPException(status_code=404, detail="用户不存在")

    user.status = UserStatusEnum(payload.status)
    write_audit(
        db,
        actor_user_id=current_user.id,
        action="UPDATE_USER_STATUS",
        target_type="user",
        target_id=str(user.id),
        request_ip=request.client.host if request.client else None,
        detail_json={"new_status": payload.status},
    )
    db.commit()
    return ApiMessage(message="状态已更新")


@router.get("/dashboard/today", dependencies=MEAL_MANAGE_DEPS)
def today_dashboard(
    current_user: Annotated[User, Depends(get_current_user)],
    db: Annotated[Session, Depends(get_db)],
    target_date: date = Query(default_factory=date.today),
):
    _ = current_user
    slots = db.scalars(select(MealSlot).where(MealSlot.meal_date == target_date)).all()
    slot_map = {s.meal_type: s for s in slots}

    counts = (
        db.execute(
            select(MealSlot.meal_type, func.count(Order.id))
            .join(Order, Order.slot_id == MealSlot.id)
            .where(MealSlot.meal_date == target_date, Order.status == OrderStatusEnum.BOOKED)
            .group_by(MealSlot.meal_type)
        )
        .all()
    )
    count_map = {meal_type: cnt for meal_type, cnt in counts}

    return DashboardStats(
        target_date=target_date,
        breakfast_count=count_map.get(MealTypeEnum.BREAKFAST, 0),
        lunch_count=count_map.get(MealTypeEnum.LUNCH, 0),
        dinner_count=count_map.get(MealTypeEnum.DINNER, 0),
        breakfast_open=slot_map.get(MealTypeEnum.BREAKFAST, MealSlot(is_open=False)).is_open,
        lunch_open=slot_map.get(MealTypeEnum.LUNCH, MealSlot(is_open=False)).is_open,
        dinner_open=slot_map.get(MealTypeEnum.DINNER, MealSlot(is_open=False)).is_open,
    )


@router.get("/meal-slots", response_model=list[AdminMealSlotOut], dependencies=MEAL_MANAGE_DEPS)
def list_meal_slots(
    current_user: Annotated[User, Depends(get_current_user)],
    db: Annotated[Session, Depends(get_db)],
    meal_date: date = Query(...),
):
    _ = current_user
    stmt = select(MealSlot).where(MealSlot.meal_date == meal_date).order_by(MealSlot.meal_type)
    return db.scalars(stmt).all()


@router.post("/meal-slots", response_model=AdminMealSlotOut, dependencies=MEAL_MANAGE_DEPS)
def create_or_update_meal_slot(
    payload: AdminMealSlotCreateRequest,
    request: Request,
    current_user: Annotated[User, Depends(get_current_user)],
    db: Annotated[Session, Depends(get_db)],
):
    meal_type_enum = MealTypeEnum(payload.meal_type)
    stmt = select(MealSlot).where(
        MealSlot.meal_date == payload.meal_date,
        MealSlot.meal_type == meal_type_enum,
    )
    slot = db.scalar(stmt)

    if slot:
        if payload.booking_deadline is not None:
            slot.booking_deadline = payload.booking_deadline
        slot.is_open = payload.is_open
        action = "UPDATE_MEAL_SLOT"
    else:
        slot = MealSlot(
            meal_date=payload.meal_date,
            meal_type=meal_type_enum,
            booking_deadline=payload.booking_deadline,
            is_open=payload.is_open,
            created_by=current_user.id,
        )
        db.add(slot)
        action = "CREATE_MEAL_SLOT"

    db.flush()
    write_audit(
        db,
        actor_user_id=current_user.id,
        action=action,
        target_type="meal_slot",
        target_id=str(slot.id),
        request_ip=request.client.host if request.client else None,
    )
    db.commit()
    db.refresh(slot)
    return slot


@router.patch("/meal-slots/{slot_id}/status", response_model=ApiMessage, dependencies=MEAL_MANAGE_DEPS)
def update_slot_status(
    slot_id: int,
    payload: AdminMealSlotStatusUpdateRequest,
    request: Request,
    current_user: Annotated[User, Depends(get_current_user)],
    db: Annotated[Session, Depends(get_db)],
):
    slot = db.get(MealSlot, slot_id)
    if not slot:
        raise HTTPException(status_code=404, detail="时段不存在")

    slot.is_open = payload.is_open
    # 清除截止时间，让订餐完全由 is_open 控制
    slot.booking_deadline = None
    write_audit(
        db,
        actor_user_id=current_user.id,
        action="UPDATE_MEAL_SLOT_STATUS",
        target_type="meal_slot",
        target_id=str(slot.id),
        request_ip=request.client.host if request.client else None,
        detail_json={"is_open": payload.is_open},
    )
    db.commit()
    return ApiMessage(message="时段状态已更新")


@router.get("/meal-packages", response_model=list[AdminMealPackageOut], dependencies=MEAL_MANAGE_DEPS)
def list_meal_packages(
    current_user: Annotated[User, Depends(get_current_user)],
    db: Annotated[Session, Depends(get_db)],
    meal_type: str | None = Query(default=None),
):
    _ = current_user
    stmt = (
        select(MealPackage)
        .options(
            joinedload(MealPackage.items),
            joinedload(MealPackage.meal_type_associations)
        )
        .where(MealPackage.is_deleted == False)
    )

    # 如果指定餐别，通过关联表过滤
    if meal_type:
        stmt = stmt.join(MealPackageMealType).where(
            MealPackageMealType.meal_type == MealTypeEnum(meal_type)
        )

    stmt = stmt.order_by(MealPackage.sort_order, MealPackage.id)
    packages = db.scalars(stmt).unique().all()
    return [_build_meal_package_out(pkg) for pkg in packages]


@router.post("/meal-packages", response_model=AdminMealPackageOut, dependencies=MEAL_MANAGE_DEPS)
def create_meal_package(
    payload: AdminMealPackageCreateRequest,
    request: Request,
    current_user: Annotated[User, Depends(get_current_user)],
    db: Annotated[Session, Depends(get_db)],
):
    # 验证餐别
    meal_type_enums = []
    for mt in payload.meal_types:
        try:
            meal_type_enums.append(MealTypeEnum(mt))
        except ValueError:
            raise HTTPException(status_code=400, detail=f"无效的餐别: {mt}")

    meal_category_enum = MealCategoryEnum(payload.meal_category)

    # 生成 package_code（不再依赖餐别，使用通用 MP 前缀）
    package_code = payload.package_code or _generate_package_code()

    # 检查 package_code 是否已存在
    existing = db.scalar(
        select(MealPackage).where(MealPackage.package_code == package_code)
    )
    if existing:
        raise HTTPException(status_code=400, detail="菜品编码已存在")

    # 获取最大 sort_order
    max_sort = db.scalar(
        select(func.max(MealPackage.sort_order)).where(MealPackage.is_deleted == False)
    )
    sort_order = (max_sort or 0) + 1

    # 创建菜品
    pkg = MealPackage(
        package_code=package_code,
        package_name=payload.package_name,
        meal_category=meal_category_enum,
        is_selectable=payload.is_selectable,
        image_url=payload.image_url,
        price=payload.price,
        unit=payload.unit,
        calories=payload.calories,
        protein_g=payload.protein_g,
        carbs_g=payload.carbs_g,
        fat_g=payload.fat_g,
        sort_order=sort_order,
    )
    db.add(pkg)
    db.flush()

    # 创建餐别关联
    for meal_type_enum in meal_type_enums:
        assoc = MealPackageMealType(
            package_id=pkg.id,
            meal_type=meal_type_enum
        )
        db.add(assoc)

    write_audit(
        db,
        actor_user_id=current_user.id,
        action="CREATE_MEAL_PACKAGE",
        target_type="meal_package",
        target_id=str(pkg.id),
        request_ip=request.client.host if request.client else None,
    )
    db.commit()
    # 重新查询带预加载，避免懒加载 N+1
    pkg = _load_package_with_relations(db, pkg.id)
    return _build_meal_package_out(pkg)


@router.post("/meal-packages/bulk-import", response_model=AdminBulkImportResult, dependencies=MEAL_MANAGE_DEPS)
async def bulk_import_meal_packages(
    file: UploadFile = File(...),
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """批量导入菜品。Excel 表头：餐别 / 菜品名称 / 分类 / 单价。
    餐别：早餐/中餐/晚餐/午晚餐；分类：普通套餐/减脂套餐/自选菜；
    单价格式：纯数字（默认单位"份"）或"金额/单位"格式（如 "5/份"、"5/个"）。
    限制：文件最大 10MB，最多 1000 行数据。"""
    if not file.filename or not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail="仅支持 .xlsx 或 .xls 文件")

    content = await file.read()
    if len(content) > settings.bulk_import_max_file_size:
        raise HTTPException(
            status_code=400,
            detail=f"文件过大，最大支持 {settings.bulk_import_max_file_size // (1024 * 1024)}MB",
        )

    try:
        wb = load_workbook(BytesIO(content), read_only=True, data_only=True)
    except Exception:
        raise HTTPException(status_code=400, detail="无效的 Excel 文件")
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise HTTPException(status_code=400, detail="Excel 文件为空")

    if len(rows) > settings.bulk_import_max_rows + 1:  # +1 for header
        raise HTTPException(
            status_code=400,
            detail=f"数据行数过多，最多支持 {settings.bulk_import_max_rows} 行",
        )

    header = [str(c or "").strip() for c in rows[0]]
    meal_type_col = next((i for i, h in enumerate(header) if "餐别" in h), None)
    name_col = next((i for i, h in enumerate(header) if "名称" in h), None)
    category_col = next((i for i, h in enumerate(header) if "分类" in h), None)
    price_col = next((i for i, h in enumerate(header) if "单价" in h or "价格" in h), None)

    if meal_type_col is None or name_col is None:
        raise HTTPException(status_code=400, detail="表头缺少「餐别」或「菜品名称」列")

    # 餐别映射：支持单餐别和组合餐别（用 / 分隔）
    meal_type_map = {
        "早餐": ["breakfast"],
        "中餐": ["lunch"],
        "晚餐": ["dinner"],
        "午餐": ["lunch"],
        "午晚餐": ["lunch", "dinner"],
        "中晚餐": ["lunch", "dinner"],
        "中餐/晚餐": ["lunch", "dinner"],
        "中餐,晚餐": ["lunch", "dinner"],
    }
    category_map = {"普通套餐": "normal", "减脂套餐": "fat_loss", "自选菜": "self_pick"}

    # Pre-load existing packages to check duplicates (by package_name)
    existing_packages = {}
    for pkg in db.scalars(
        select(MealPackage).where(MealPackage.is_deleted == False)
    ).all():
        existing_packages[pkg.package_name] = pkg.id

    created = 0
    skipped = 0
    errors = []

    try:
        for idx, row in enumerate(rows[1:], start=2):
            if not row or all(c is None or str(c).strip() == "" for c in row):
                continue

            meal_type_raw = str(row[meal_type_col]).strip() if row[meal_type_col] else None
            package_name = str(row[name_col]).strip() if row[name_col] else None
            category_raw = str(row[category_col]).strip() if category_col is not None and row[category_col] else "普通套餐"
            price_raw = row[price_col] if price_col is not None and row[price_col] else 0

            if not meal_type_raw or not package_name:
                errors.append(f"第 {idx} 行：餐别或菜品名称为空")
                continue

            meal_types = meal_type_map.get(meal_type_raw)
            if not meal_types:
                errors.append(f"第 {idx} 行：餐别「{meal_type_raw}」无效（支持：早餐/中餐/晚餐/午晚餐)")
                continue

            category = category_map.get(category_raw, "normal")

            # 解析单价：支持 "5"、"5.5"、"5/份"、"5/个" 等格式
            unit = "份"
            price_str = str(price_raw).strip() if price_raw is not None else "0"
            if "/" in price_str:
                parts = price_str.split("/", 1)
                price_str = parts[0].strip()
                parsed_unit = parts[1].strip()
                if parsed_unit:
                    if len(parsed_unit) > 16:
                        errors.append(f"第 {idx} 行：单位「{parsed_unit}」过长（最多 16 个字符）")
                        continue
                    unit = parsed_unit

            try:
                price = float(price_str)
                if price < 0:
                    errors.append(f"第 {idx} 行：单价不能为负数")
                    continue
            except (ValueError, TypeError):
                errors.append(f"第 {idx} 行：单价「{price_raw}」无效")
                continue

            # Check duplicate (by package_name only)
            if package_name in existing_packages:
                skipped += 1
                continue

            package_code = _generate_package_code()

            max_sort = db.scalar(
                select(func.max(MealPackage.sort_order)).where(MealPackage.is_deleted == False)
            )
            sort_order = (max_sort or 0) + 1

            pkg = MealPackage(
                package_code=package_code,
                package_name=package_name,
                meal_category=MealCategoryEnum(category),
                is_selectable=True,
                price=price,
                unit=unit,
                sort_order=sort_order,
            )
            db.add(pkg)
            db.flush()

            # 创建餐别关联
            for mt in meal_types:
                assoc = MealPackageMealType(
                    package_id=pkg.id,
                    meal_type=MealTypeEnum(mt)
                )
                db.add(assoc)

            existing_packages[package_name] = pkg.id
            created += 1

        db.commit()
    except IntegrityError as e:
        db.rollback()
        error_msg = str(e.orig) if e.orig else str(e)
        raise HTTPException(status_code=409, detail=f"批量导入失败：数据冲突 - {error_msg}")
    except ValueError as e:
        db.rollback()
        raise HTTPException(status_code=400, detail=f"批量导入失败：数据格式错误 - {str(e)}")
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"批量导入失败：{str(e)}")

    write_audit(
        db,
        actor_user_id=current_user.id,
        action="BULK_IMPORT_MEAL_PACKAGES",
        target_type="meal_package",
        target_id="bulk",
        detail_json={"created": created, "skipped": skipped, "error_count": len(errors)},
    )
    db.commit()

    return AdminBulkImportResult(created=created, skipped=skipped, errors=errors)


@router.patch("/meal-packages/{package_id}", response_model=AdminMealPackageOut, dependencies=MEAL_MANAGE_DEPS)
def update_meal_package(
    package_id: int,
    payload: AdminMealPackageUpdateRequest,
    request: Request,
    current_user: Annotated[User, Depends(get_current_user)],
    db: Annotated[Session, Depends(get_db)],
):
    # 预加载 meal_type_associations，便于后续 diff 更新
    pkg = _load_package_with_relations(db, package_id)
    if not pkg or pkg.is_deleted:
        raise HTTPException(status_code=404, detail="菜品不存在")

    # 更新餐别关联（diff 模式：只删除待移除的、只添加新增的）
    if payload.meal_types is not None:
        # 验证餐别
        new_meal_types: set[MealTypeEnum] = set()
        for mt in payload.meal_types:
            try:
                new_meal_types.add(MealTypeEnum(mt))
            except ValueError:
                raise HTTPException(status_code=400, detail=f"无效的餐别: {mt}")

        # 计算 diff
        current_assocs = {assoc.meal_type: assoc for assoc in pkg.meal_type_associations}
        current_meal_types = set(current_assocs.keys())

        to_remove = current_meal_types - new_meal_types
        to_add = new_meal_types - current_meal_types

        # 只删除待移除的关联
        for mt in to_remove:
            db.delete(current_assocs[mt])

        # 只添加新增的关联
        for mt in to_add:
            db.add(MealPackageMealType(package_id=pkg.id, meal_type=mt))

    if payload.package_name is not None:
        pkg.package_name = payload.package_name
    if payload.meal_category is not None:
        pkg.meal_category = MealCategoryEnum(payload.meal_category)
    if payload.image_url is not None:
        pkg.image_url = payload.image_url
    if payload.price is not None:
        pkg.price = payload.price
    if payload.unit is not None:
        pkg.unit = payload.unit
    if payload.calories is not None:
        pkg.calories = payload.calories
    if payload.protein_g is not None:
        pkg.protein_g = payload.protein_g
    if payload.carbs_g is not None:
        pkg.carbs_g = payload.carbs_g
    if payload.fat_g is not None:
        pkg.fat_g = payload.fat_g
    if payload.is_selectable is not None:
        pkg.is_selectable = payload.is_selectable

    write_audit(
        db,
        actor_user_id=current_user.id,
        action="UPDATE_MEAL_PACKAGE",
        target_type="meal_package",
        target_id=str(pkg.id),
        request_ip=request.client.host if request.client else None,
    )
    db.commit()
    # 重新查询带预加载，避免懒加载 N+1
    pkg = _load_package_with_relations(db, pkg.id)
    return _build_meal_package_out(pkg)


@router.post("/uploads/meal-image", dependencies=MEAL_MANAGE_DEPS)
async def upload_meal_image(
    file: UploadFile = File(...),
    current_user: User = Depends(get_current_user),
):
    _ = current_user
    if not file.filename:
        raise HTTPException(status_code=400, detail="文件名为空")

    ext = Path(file.filename).suffix.lower()
    if ext not in {".jpg", ".jpeg", ".png", ".webp"}:
        raise HTTPException(status_code=400, detail="仅支持 jpg/png/webp 图片")

    upload_dir = Path(settings.upload_dir) / "meal_images"
    upload_dir.mkdir(parents=True, exist_ok=True)

    filename = f"{uuid4().hex}{ext}"
    file_path = upload_dir / filename
    content = await file.read()
    file_path.write_bytes(content)

    image_url = f"/uploads/meal_images/{filename}"
    return {"image_url": image_url}


@router.delete("/meal-packages/{package_id}", response_model=ApiMessage, dependencies=MEAL_MANAGE_DEPS)
def delete_meal_package(
    package_id: int,
    request: Request,
    current_user: Annotated[User, Depends(get_current_user)],
    db: Annotated[Session, Depends(get_db)],
):
    # 预加载关联，审计日志需要 meal_types
    pkg = _load_package_with_relations(db, package_id)
    if not pkg or pkg.is_deleted:
        raise HTTPException(status_code=404, detail="菜品不存在")

    future_orders = db.scalars(
        select(Order)
        .join(MealSlot, Order.slot_id == MealSlot.id)
        .where(
            Order.package_id == package_id,
            Order.status == OrderStatusEnum.BOOKED,
            MealSlot.meal_date >= date.today(),
        )
    ).all()

    cancelled_order_ids = []
    trimmed_order_ids = []

    for order in future_orders:
        if order.meal_category == MealCategoryEnum.SELF_PICK:
            db.execute(select(OrderItem).where(OrderItem.order_id == order.id, OrderItem.item_name == pkg.package_name))
            items = db.scalars(select(OrderItem).where(OrderItem.order_id == order.id)).all()
            if len(items) <= 1:
                order.status = OrderStatusEnum.CANCELLED
                order.cancelled_at = datetime.utcnow()
                cancelled_order_ids.append(order.id)
            else:
                db.execute(
                    select(OrderItem).where(OrderItem.order_id == order.id, OrderItem.item_name == pkg.package_name)
                )
                for item in db.scalars(
                    select(OrderItem).where(OrderItem.order_id == order.id, OrderItem.item_name == pkg.package_name)
                ).all():
                    db.delete(item)
                trimmed_order_ids.append(order.id)
        else:
            order.status = OrderStatusEnum.CANCELLED
            order.cancelled_at = datetime.utcnow()
            cancelled_order_ids.append(order.id)

    pkg.is_deleted = True
    write_audit(
        db,
        actor_user_id=current_user.id,
        action="DELETE_MEAL_PACKAGE",
        target_type="meal_package",
        target_id=str(pkg.id),
        request_ip=request.client.host if request.client else None,
        detail_json={
            "meal_types": [mt.value for mt in pkg.meal_types],
            "package_name": pkg.package_name,
            "cancelled_order_ids": cancelled_order_ids,
            "trimmed_order_ids": trimmed_order_ids,
        },
    )
    db.commit()
    return ApiMessage(message="菜品已删除")
